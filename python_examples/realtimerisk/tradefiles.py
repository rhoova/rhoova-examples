import os
from openpyxl import load_workbook

def find_pos_def(excelfile,position):
    filename=excelfile
    wb = load_workbook(filename)  # load work book
    ws = wb.active  # get active work sheet
    for row in ws.rows:  # iterate over each row
        for cell in row:  # iterate over each cell in the row
            if cell.value == position:  # find "part number"
                return ws.cell(row=cell.row, column=cell.col_idx).value  # get the value of the  cell
    return None

def findFiles(strings, dir, subDirs, fileContent, fileExtensions):
    filesInDir = []
    foundFiles = {}
    filesFound = 0
    filesname = [name for name in os.listdir(dir) if name !='.DS_Store']
    #if not subDirs:
    for filename in filesname :
        if os.path.isfile(os.path.join(dir, filename).replace("\\", "/")):
            filesInDir.append(os.path.join(dir, filename).replace("\\", "/"))
    #else:
    #    for root, subdirs, files in os.walk(dir):
    #        for f in files:
    #            if not os.path.isdir(os.path.join(root, f).replace("\\", "/")):
    #                filesInDir.append(os.path.join(root, f).replace("\\", "/"))
    if filesInDir:
        for file in filesInDir:
            #print("Current file: "+file)
            filename, extension = os.path.splitext(file)
            if fileExtensions:
                fileText = extension
            else:
                fileText = os.path.basename(filename).lower()
                if fileContent:
                    fileText +=  getFileContent(file).lower()
            
            for string in strings:
                pos_def=find_pos_def(file,string)
                if pos_def is not None:
                    #print(file,pos_def)
                    foundFiles[pos_def]=file
                    #foundFiles.append(file_dict)
                #if string in fileText:
                #    print(string,file)
                #    foundFiles.append(file)
                    #filesFound += 1
                    #break
    return foundFiles

def getFileContent(filename):
    supportedTypes = [".xls", ".xlsx"]
    if filename.partition(".")[2] in supportedTypes:
        if filename.endswith(".xls"):
            content = ""
            with openpyxl.load_workbook(filename) as pdf:
                for x in range(0, len(pdf.pages)):
                    page = pdf.pages[x]
                    content = content + page.extract_text()
            return content
        elif filename.endswith(".xlsx"):
            with openpyxl.load_workbook(filename, 'r') as f:
                content = ""
                lines = f.readlines()
                for x in lines:
                    content = content + x
            f.close()
            return content
    else:
        return ""